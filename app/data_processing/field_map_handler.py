import pandas as pd
import openpyxl
import os
import re

class FieldMapHandler:
    def __init__(self, field_map_path, tube_ids_path):
        self.field_map_path = field_map_path
        self.tube_ids_path = tube_ids_path
        self.metadata_map = {} # Tube -> {Treatment, Genotype, Rng, Col}
        
    def load_data(self):
        """Loads and merges tube, plot, and field map data."""
        if not os.path.exists(self.field_map_path):
            print(f"Field Map not found at: {self.field_map_path}")
            return {}
        if not os.path.exists(self.tube_ids_path):
            print(f"Tube IDs not found at: {self.tube_ids_path}")
            return {}
            
        # 1. Load Tube Mapping (Tube -> Rng, Col)
        try:
            tube_df = pd.read_csv(self.tube_ids_path)
            # Ensure columns exist
            required_cols = ['Tube', 'Rng', 'Col']
            if not all(col in tube_df.columns for col in required_cols):
                print(f"TubeIDS.csv missing required columns: {required_cols}")
                return {}
        except Exception as e:
            print(f"Error reading TubeIDS.csv: {e}")
            return {}
        
        # 2. Load Field Map (Excel Colors & Text)
        try:
            # Use data_only=False to preserve cell styles/colors
            wb = openpyxl.load_workbook(self.field_map_path, data_only=False)
            sheet = wb.active
        except Exception as e:
            print(f"Error reading FieldMap.xlsx: {e}")
            return {}
        
        # Map to store tube -> metadata (found by searching for tube IDs in Excel)
        tube_to_metadata = {}
        
        print("Scanning Field Map for tube IDs and colors...")
        
        def get_color_hex(cell):
            """Extract color as hex string from cell"""
            if not cell.fill or cell.fill.patternType != 'solid':
                return None
            fg_color = cell.fill.fgColor
            if hasattr(fg_color, 'rgb') and fg_color.rgb:
                try:
                    rgb = fg_color.rgb
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        return rgb.upper()
                except:
                    pass
            return None
        
        def map_color_to_treatment(color_hex):
            """Map color hex code to treatment name based on actual FieldMap.xlsx colors"""
            if not color_hex:
                return "Unknown"
            
            # Actual colors found in FieldMap.xlsx (ARGB format, 8 chars):
            # FF00B0F0 - Light blue/cyan (Water/Control)
            # FFFF0000 - Red (Drought)
            # FF00B050 - Green (likely another treatment)
            # FFFF00 - Yellow (likely another treatment)
            
            color_upper = color_hex.upper()
            
            # Normalize to 8-char ARGB format if needed (add FF prefix if 6 chars)
            if len(color_upper) == 6:
                color_upper = "FF" + color_upper
            
            # Match exact ARGB codes (8 characters)
            if color_upper == "FF00B0F0":
                return "Water (Control)"
            elif color_upper == "FFFF0000":
                return "Drought"
            elif color_upper == "FF00B050":
                return "Treatment 3"  # Green - may need to be renamed based on actual treatment
            elif color_upper == "FFFFFF00":
                return "Treatment 4"  # Yellow - may need to be renamed based on actual treatment
            else:
                return "Unknown"
        
        # Scan Excel for tube IDs (pattern: T{number} or just {number})
        for row in sheet.iter_rows(min_row=14, max_row=sheet.max_row):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                
                value_str = str(value).strip()
                
                # Look for tube ID pattern: T{number} (e.g., "T17-P4" or "T17")
                match = re.search(r'T(\d+)', value_str)
                if match:
                    tube_num = int(match.group(1))
                    
                    # Extract color
                    color_hex = get_color_hex(cell)
                    treatment = map_color_to_treatment(color_hex)
                    
                    # Extract genotype from cell value
                    # Genotype might be in the format like "SC-G4-18585019" or "TC-G3-18585041"
                    # or might be part of the tube ID string
                    genotype = "Unknown"
                    
                    # Check if value contains genotype codes (SC-G, TC-G, etc.)
                    genotype_match = re.search(r'([ST]C-G\d+-\d+)', value_str)
                    if genotype_match:
                        genotype = genotype_match.group(1)
                    else:
                        # If no explicit genotype, use the full value as genotype
                        # (excluding the tube ID part)
                        genotype = value_str
                    
                    tube_to_metadata[tube_num] = {
                        "Treatment": treatment,
                        "Genotype": genotype,
                        "Color": color_hex
                    }
        
        wb.close()

        # 3. Merge Tube Map with Field Data
        for _, row in tube_df.iterrows():
            try:
                tube = int(row['Tube'])
                rng = int(row['Rng'])
                col = int(row['Col'])
                
                # Look up metadata from Excel search
                meta = tube_to_metadata.get(tube, {
                    "Treatment": "Unknown",
                    "Genotype": "Unknown"
                })

                self.metadata_map[tube] = {
                    "Rng": rng,
                    "Col": col,
                    "Treatment": meta["Treatment"],
                    "Genotype": meta["Genotype"]
                }
            except ValueError:
                continue
            
        return self.metadata_map

    def get_tube_metadata(self, tube_id):
        return self.metadata_map.get(int(tube_id), {})

